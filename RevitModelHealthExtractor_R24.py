# CHANGELOG:
# - [2026-06-16] Refactored script into modular functions for better readability.
# - [2026-06-16] Added Boolean toggles via Dynamo IN[...] nodes to let user control which modules to run.
# - [2026-06-16] Maintained 'bulletproof' logic for Group properties bypassing Python.NET bugs.

import sys
import clr
import System
import os
from datetime import datetime

# =======================================================
# INIT REVIT API FOR DYNAMO
# =======================================================
clr.AddReference('RevitAPI')
from Autodesk.Revit.DB import *

clr.AddReference("RevitServices")
import RevitServices
from RevitServices.Persistence import DocumentManager

doc = DocumentManager.Instance.CurrentDBDocument

# =======================================================
# USER INPUTS & UI
# =======================================================
try:
    ENABLE_SCRIPT = IN[0] if len(IN) > 0 and IN[0] is not None else False
except NameError:
    ENABLE_SCRIPT = False

OUTPUT_FILEPATH = None
USER_NOTE = ""

if ENABLE_SCRIPT:
    clr.AddReference("System.Windows.Forms")
    from System.Windows.Forms import SaveFileDialog, DialogResult
    
    dialog = SaveFileDialog()
    dialog.Filter = "Excel Files (*.xlsx)|*.xlsx"
    dialog.Title = "Save Model Health Report"
    
    if dialog.ShowDialog() == DialogResult.OK:
        OUTPUT_FILEPATH = dialog.FileName
        
        # Pop up Note Input
        from System.Drawing import Point, Size
        from System.Windows.Forms import Form, TextBox, Button, Label, FormBorderStyle, FormStartPosition
        
        form = Form()
        form.Text = "Extraction Note"
        form.Width = 400
        form.Height = 160
        form.FormBorderStyle = FormBorderStyle.FixedDialog
        form.StartPosition = FormStartPosition.CenterScreen
        form.MaximizeBox = False
        form.MinimizeBox = False
        
        lbl = Label()
        lbl.Text = "Enter Note / Phase (e.g., 50% DD, Pre-Submission):"
        lbl.Location = Point(10, 10)
        lbl.AutoSize = True
        
        txt = TextBox()
        txt.Location = Point(10, 35)
        txt.Width = 360
        
        btn_ok = Button()
        btn_ok.Text = "OK"
        btn_ok.Location = Point(200, 75)
        btn_ok.DialogResult = DialogResult.OK
        
        btn_cancel = Button()
        btn_cancel.Text = "Cancel"
        btn_cancel.Location = Point(290, 75)
        btn_cancel.DialogResult = DialogResult.Cancel
        
        form.AcceptButton = btn_ok
        form.CancelButton = btn_cancel
        form.Controls.Add(lbl)
        form.Controls.Add(txt)
        form.Controls.Add(btn_ok)
        form.Controls.Add(btn_cancel)
        
        if form.ShowDialog() == DialogResult.OK:
            USER_NOTE = txt.Text
        else:
            USER_NOTE = ""

ENABLE_WARNINGS = True
ENABLE_GROUPS = True
ENABLE_CAD_LINKS = True
ENABLE_VIEWS_SHEETS = True
ENABLE_LINE_STYLES = True
ENABLE_FAMILIES = True
ENABLE_WORKSETS_LINKS = True

# =======================================================
# CORE HELPER FUNCTIONS
# =======================================================
def get_file_size_mb(doc):
    try:
        from Autodesk.Revit.DB import ModelPathUtils
        if doc.IsModelInCloud:
            return "Cloud Model"
        
        path_to_check = doc.PathName
        
        # If workshared (Create Local), attempt to find Central Model Size
        if doc.IsWorkshared:
            try:
                central_path = doc.GetWorksharingCentralModelPath()
                if central_path:
                    visible_path = ModelPathUtils.ConvertModelPathToUserVisiblePath(central_path)
                    if visible_path:
                        if visible_path.startswith("RSN://"):
                            return "Revit Server"
                        elif os.path.exists(visible_path):
                            # Prioritize measuring the central file size!
                            path_to_check = visible_path
            except Exception:
                pass
                
        if not path_to_check:
            return "Detached/Unsaved"
            
        if path_to_check.startswith("RSN://"):
            return "Revit Server"
            
        if os.path.exists(path_to_check):
            size_bytes = os.path.getsize(path_to_check)
            return round(size_bytes / (1024 * 1024), 2)
            
    except Exception:
        pass
    return "Unknown Path"

def get_creator(elem):
    if doc.IsWorkshared and elem is not None:
        try:
            info = WorksharingUtils.GetWorksharingTooltipInfo(doc, elem.Id)
            return info.Creator if info.Creator else ""
        except Exception:
            return ""
    return ""

def get_last_changed_by(elem):
    if doc.IsWorkshared and elem is not None:
        try:
            info = WorksharingUtils.GetWorksharingTooltipInfo(doc, elem.Id)
            return info.LastChangedBy if info.LastChangedBy else ""
        except Exception:
            return ""
    return ""

def count_faces(element):
    faces = 0
    try:
        opt = Options()
        opt.DetailLevel = ViewDetailLevel.Fine
        geom = element.get_Geometry(opt)
        if geom:
            for g_obj in geom:
                if hasattr(g_obj, "Faces") and g_obj.Faces.Size > 0:
                    faces += g_obj.Faces.Size
                elif hasattr(g_obj, "GetInstanceGeometry"):
                    inst_geom = g_obj.GetInstanceGeometry()
                    if inst_geom:
                        for i_obj in inst_geom:
                            if hasattr(i_obj, "Faces") and i_obj.Faces.Size > 0:
                                faces += i_obj.Faces.Size
    except Exception:
        pass
    return faces

def safe_get_name(elem):
    try:
        return Element.Name.__get__(elem)
    except Exception:
        pass
    try:
        if hasattr(elem, "Name"):
            return elem.Name
    except Exception:
        pass
    try:
        param = elem.get_Parameter(BuiltInParameter.SYMBOL_NAME_PARAM)
        if param: return param.AsString()
    except Exception:
        pass
    return "Unknown Name"

def safe_get_category_name(elem):
    try:
        if hasattr(elem, "Category") and elem.Category:
            return elem.Category.Name
    except Exception:
        pass
    return "Unknown Category"

# =======================================================
# EXTRACTION MODULES
# =======================================================

def extract_project_info():
    proj_info = doc.ProjectInformation
    proj_name = safe_get_name(proj_info) if proj_info else "Unknown"
    proj_number = proj_info.Number if proj_info and hasattr(proj_info, "Number") else "Unknown"
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    stamp_headers = ["Date Extracted", "Project Name", "Project Number"]
    stamp_data = [timestamp, proj_name, proj_number]
    return stamp_headers, stamp_data, timestamp

def extract_warnings(stamp_data):
    if not ENABLE_WARNINGS: return []
    warnings = doc.GetWarnings()
    data = []
    for w in warnings:
        desc = w.GetDescriptionText()
        severity = str(w.GetSeverity())
        failing_elems = ", ".join([str(id.IntegerValue) for id in w.GetFailingElements()])
        data.append(stamp_data + [desc, severity, failing_elems])
    return data

def extract_inplace_families(stamp_data):
    if not ENABLE_FAMILIES: return []
    data = []
    all_family_instances = FilteredElementCollector(doc).OfClass(FamilyInstance).ToElements()
    for fi in all_family_instances:
        if fi.Symbol and fi.Symbol.Family and fi.Symbol.Family.IsInPlace:
            cat_name = safe_get_category_name(fi)
            fam_name = safe_get_name(fi.Symbol.Family)
            creator = get_creator(fi)
            last_changed_by = get_last_changed_by(fi)
            data.append(stamp_data + [fi.Id.IntegerValue, cat_name, fam_name, creator, last_changed_by])
    return data

def extract_views_and_sheets(stamp_data):
    if not ENABLE_VIEWS_SHEETS: return []
    data = []
    all_views = FilteredElementCollector(doc).OfClass(View).ToElements()
    viewports = FilteredElementCollector(doc).OfClass(Viewport).ToElements()
    placed_view_ids = set([vp.ViewId.IntegerValue for vp in viewports])
    
    for v in all_views:
        if not v.IsTemplate:
            is_sheet = isinstance(v, ViewSheet)
            v_type = "Sheet" if is_sheet else "View"
            v_id = v.Id.IntegerValue
            v_name = safe_get_name(v)
            
            template_name = "<None>"
            if v.ViewTemplateId != ElementId.InvalidElementId:
                template_elem = doc.GetElement(v.ViewTemplateId)
                if template_elem:
                    template_name = safe_get_name(template_elem)
            
            is_placed = "N/A"
            if not is_sheet:
                is_placed = "Yes" if (v_id in placed_view_ids) else "No"
                
            creator = get_creator(v)
            last_changed_by = get_last_changed_by(v)
            data.append(stamp_data + [v_type, v_id, v_name, template_name, is_placed, creator, last_changed_by])
    return data

def extract_cad_links(stamp_data):
    if not ENABLE_CAD_LINKS: return []
    data = []
    import_instances = FilteredElementCollector(doc).OfClass(ImportInstance).ToElements()
    
    for imp in import_instances:
        is_linked = imp.IsLinked
        cad_name = "Unknown"
        imp_type = doc.GetElement(imp.GetTypeId())
        if imp_type:
            cad_name = safe_get_name(imp_type)
            
        if cad_name in ["Unknown", "Unknown Name", ""]:
            cad_name = safe_get_category_name(imp)

        creator = get_creator(imp)
        last_changed_by = get_last_changed_by(imp)
        link_type_str = "Linked" if is_linked else "Imported"
        data.append(stamp_data + [imp.Id.IntegerValue, cad_name, link_type_str, creator, last_changed_by])
    return data

def extract_worksets(stamp_data):
    if not ENABLE_WORKSETS_LINKS: return []
    data = []
    if doc.IsWorkshared:
        worksets = FilteredWorksetCollector(doc).OfKind(WorksetKind.UserWorkset)
        for ws in worksets:
            data.append(stamp_data + [ws.Id.IntegerValue, ws.Name, str(ws.IsDefaultWorkset)])
    else:
        data.append(stamp_data + ["N/A", "Model is not workshared", "N/A"])
    return data

def extract_revit_links(stamp_data):
    if not ENABLE_WORKSETS_LINKS: return []
    data = []
    try:
        rvt_link_instances = FilteredElementCollector(doc).OfClass(RevitLinkInstance).ToElements()
        for link_inst in rvt_link_instances:
            link_name = safe_get_name(link_inst)
            is_pinned = "Yes" if link_inst.Pinned else "No"
            workset_id = link_inst.WorksetId.IntegerValue
            creator = get_creator(link_inst)
            last_changed_by = get_last_changed_by(link_inst)
            
            loaded_status = "Unknown"
            ref_type = "Unknown"
            link_type = doc.GetElement(link_inst.GetTypeId())
            if link_type:
                try:
                    loaded_status = str(link_type.GetLinkedFileStatus())
                    ref_type = str(link_type.AttachmentType)
                except Exception:
                    pass
            data.append(stamp_data + [link_name, loaded_status, ref_type, is_pinned, workset_id, creator, last_changed_by])
    except Exception:
        pass
    return data

def extract_groups(stamp_data):
    if not ENABLE_GROUPS: return []
    data = []
    try:
        group_types = FilteredElementCollector(doc).OfClass(GroupType).ToElements()
        for gtype in group_types:
            g_name = safe_get_name(gtype)
            g_category = safe_get_category_name(gtype)
            creator = get_creator(gtype)
            last_changed_by = get_last_changed_by(gtype)
            
            placement_count = 0
            member_count = 0
            try:
                placements = gtype.Groups
                if placements is not None:
                    placement_count = placements.Size
                    if placement_count > 0:
                        it = placements.GetEnumerator()
                        if it.MoveNext() and it.Current:
                            member_count = len(it.Current.GetMemberIds())
            except Exception:
                placement_count = -1
                
            data.append(stamp_data + [g_name, g_category, placement_count, member_count, creator, last_changed_by])
    except Exception:
        pass
    return data

def extract_linestyles_patterns(stamp_data):
    if not ENABLE_LINE_STYLES: return []
    data = []
    suspicious_keywords = ["autocad", "import", ".dwg", "solid - ", "solid_"]
    try:
        lines_cat = Category.GetCategory(doc, BuiltInCategory.OST_Lines)
        if lines_cat:
            for subcat in lines_cat.SubCategories:
                name = safe_get_name(subcat)
                is_suspicious = "Yes" if any(kw in name.lower() for kw in suspicious_keywords) else "No"
                creator = get_creator(subcat)
                last_changed_by = get_last_changed_by(subcat)
                gs = subcat.GetGraphicsStyle(GraphicsStyleType.Projection)
                if gs: 
                    creator = get_creator(gs)
                    last_changed_by = get_last_changed_by(gs)
                data.append(stamp_data + [name, "Line Style", is_suspicious, creator, last_changed_by])
                
        fill_patterns = FilteredElementCollector(doc).OfClass(FillPatternElement).ToElements()
        for fp in fill_patterns:
            name = safe_get_name(fp)
            is_suspicious = "Yes" if any(kw in name.lower() for kw in suspicious_keywords) else "No"
            creator = get_creator(fp)
            last_changed_by = get_last_changed_by(fp)
            data.append(stamp_data + [name, "Fill Pattern", is_suspicious, creator, last_changed_by])
    except Exception:
        pass
    return data

def extract_loadable_families(stamp_data):
    if not ENABLE_FAMILIES: return []
    data = []
    try:
        all_families = FilteredElementCollector(doc).OfClass(Family).ToElements()
        all_instances = FilteredElementCollector(doc).WhereElementIsNotElementType().ToElements()
        
        instance_counts = {}
        family_sample_instances = {}
        for inst in all_instances:
            try:
                type_id = inst.GetTypeId()
                if type_id != ElementId.InvalidElementId:
                    elem_type = doc.GetElement(type_id)
                    if hasattr(elem_type, "Family") and elem_type.Family:
                        fam_id = elem_type.Family.Id.IntegerValue
                        instance_counts[fam_id] = instance_counts.get(fam_id, 0) + 1
                        if fam_id not in family_sample_instances:
                            family_sample_instances[fam_id] = inst
            except Exception:
                pass

        fam_list = []
        for fam in all_families:
            if not fam.IsInPlace:
                f_name = safe_get_name(fam)
                f_cat = fam.FamilyCategory.Name if fam.FamilyCategory else "Unknown Category"
                creator = get_creator(fam)
                last_changed_by = get_last_changed_by(fam)
                fam_id = fam.Id.IntegerValue
                p_count = instance_counts.get(fam_id, 0)
                
                face_count = 0
                risk_level = "Unknown"
                try:
                    if fam_id in family_sample_instances:
                        sample_inst = family_sample_instances[fam_id]
                        face_count = count_faces(sample_inst)
                    else:
                        symbol_ids = fam.GetFamilySymbolIds()
                        if symbol_ids.Count > 0:
                            sym = doc.GetElement(symbol_ids[0])
                            face_count = count_faces(sym)
                            
                    if face_count > 5000: risk_level = "🔴 High"
                    elif face_count > 1500: risk_level = "🟡 Medium"
                    else: risk_level = "🟢 Low"
                except Exception:
                    pass
                
                fam_list.append([f_cat, f_name, p_count, face_count, risk_level, creator, last_changed_by])
                
        fam_list.sort(key=lambda x: (x[0], x[1]))
        for row in fam_list:
            data.append(stamp_data + row)
    except Exception:
        pass
    return data

# =======================================================
# MAIN EXECUTION
# =======================================================
stamp_headers, stamp_data, timestamp = extract_project_info()

d_warnings = extract_warnings(stamp_data)
d_inplace = extract_inplace_families(stamp_data)
d_views = extract_views_and_sheets(stamp_data)
d_cad = extract_cad_links(stamp_data)
d_worksets = extract_worksets(stamp_data)
d_rvt = extract_revit_links(stamp_data)
d_groups = extract_groups(stamp_data)
d_lines = extract_linestyles_patterns(stamp_data)
d_loadable = extract_loadable_families(stamp_data)

# Calculate summary
file_size = get_file_size_mb(doc)
total_warnings = len(doc.GetWarnings()) if ENABLE_WARNINGS else "Skipped"

# Create datasets array for Excel writing
datasets = []

# Prepend Headers
if ENABLE_WARNINGS: datasets.append(("Warnings", [stamp_headers + ["Warning Description", "Severity", "Failing Elements (IDs)"]] + d_warnings))
if ENABLE_FAMILIES: datasets.append(("InPlaceFamilies", [stamp_headers + ["Element ID", "Category", "Family Name", "Creator (Owner)", "LastChangedBy"]] + d_inplace))
if ENABLE_VIEWS_SHEETS: datasets.append(("ViewsAndSheets", [stamp_headers + ["Item Type", "Item ID", "Name", "View Template Applied", "Is Placed on Sheet", "Creator", "LastChangedBy"]] + d_views))
if ENABLE_CAD_LINKS: datasets.append(("CADLinks", [stamp_headers + ["Element ID", "CAD File Name", "Type (Linked/Imported)", "Creator", "LastChangedBy"]] + d_cad))
if ENABLE_WORKSETS_LINKS: datasets.append(("Worksets", [stamp_headers + ["Workset ID", "Workset Name", "Is Default"]] + d_worksets))
if ENABLE_WORKSETS_LINKS: datasets.append(("RevitLinks", [stamp_headers + ["Link Name", "Loaded Status", "Reference Type", "Pinned Status", "Workset ID", "Creator", "LastChangedBy"]] + d_rvt))
if ENABLE_GROUPS: datasets.append(("Groups", [stamp_headers + ["Group Name", "Group Category", "Placement Count", "Member Count", "Creator", "LastChangedBy"]] + d_groups))
if ENABLE_LINE_STYLES: datasets.append(("LineStylesAndPatterns", [stamp_headers + ["Item Name", "Type", "Is Suspicious", "Creator", "LastChangedBy"]] + d_lines))
if ENABLE_FAMILIES: datasets.append(("LoadableFamilies", [stamp_headers + ["Category", "Family Name", "Placement Count", "Polygon/Face Count", "Risk Level", "Creator", "LastChangedBy"]] + d_loadable))

# Summary
summary_headers = stamp_headers + ["Note", "File Size (MB)", "Total Warnings", "In-Place Count", "Views Count", "Sheets Count", "CAD Links", "Worksets", "Revit Links", "Groups", "Line Styles/Patterns", "Families"]
summary_values = stamp_data + [
    USER_NOTE, file_size, total_warnings, 
    len(d_inplace), len(d_views), sum(1 for v in d_views if v[3] == "Sheet"), 
    len(d_cad), len(d_worksets), len(d_rvt), len(d_groups), len(d_lines), len(d_loadable)
]
datasets.insert(0, ("Summary", [summary_headers, summary_values]))

# =======================================================
# EXCEL EXPORT (OPENPYXL)
# =======================================================
if OUTPUT_FILEPATH and str(OUTPUT_FILEPATH).lower().endswith(".xlsx"):
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
        
        wb = None
        if os.path.exists(OUTPUT_FILEPATH):
            try:
                wb = load_workbook(OUTPUT_FILEPATH)
            except Exception:
                wb = Workbook()
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
            
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        group_fill = PatternFill("solid", fgColor="E0E0E0")
        group_font = Font(bold=True)
        group_alignment = Alignment(horizontal="centerContinuous", vertical="center")
        
        cat_fill = PatternFill("solid", fgColor="F2F2F2")
        cat_font = Font(italic=True, bold=True)

        for name, data in datasets:
            if not data or len(data) <= 1: continue
            
            if name in wb.sheetnames: ws = wb[name]
            else: ws = wb.create_sheet(title=name)
            
            start_row = ws.max_row
            if start_row == 1 and not ws.cell(row=1, column=1).value: start_row = 0
            
            is_new_sheet = (start_row == 0)
            num_cols = len(data[0])
            current_row = start_row + 1
            
            # Header
            if is_new_sheet:
                for c_idx, value in enumerate(data[0], 1):
                    cell = ws.cell(row=current_row, column=c_idx, value=str(value) if value is not None else "")
                    cell.font, cell.fill, cell.border = header_font, header_fill, border
                current_row += 1
                
            # Date Group Header
            group_header_text = "--- 🕒 รอบการอัปเดต: " + str(timestamp) + " ---"
            for c_idx in range(1, num_cols + 1):
                cell = ws.cell(row=current_row, column=c_idx)
                cell.fill, cell.alignment = group_fill, group_alignment
            ws.cell(row=current_row, column=1, value=group_header_text).font = group_font
            
            # Data Rows with internal categorization
            rows_to_write = data[1:]
            needs_category = False
            cat_idx = -1
            group_prefix = "📂 หมวดหมู่:"
            
            if name == "Groups": needs_category, cat_idx = True, 4
            elif name == "LoadableFamilies": needs_category, cat_idx = True, 3
            elif name == "InPlaceFamilies": needs_category, cat_idx = True, 4
            elif name == "ViewsAndSheets": needs_category, cat_idx, group_prefix = True, 8, "👤 ผู้สร้าง/เจ้าของ:"
            elif name == "CADLinks": needs_category, cat_idx, group_prefix = True, 6, "👤 ผู้นำเข้า:"
            elif name == "LineStylesAndPatterns": needs_category, cat_idx, group_prefix = True, 6, "👤 ผู้นำเข้า:"
            
            if needs_category:
                if name == "InPlaceFamilies": rows_to_write.sort(key=lambda x: (str(x[4]), str(x[5])))
                elif name == "LoadableFamilies": rows_to_write.sort(key=lambda x: (str(x[3]), str(x[4])))
                else: rows_to_write.sort(key=lambda x: str(x[cat_idx]))
                
            row_pointer = current_row + 1
            current_category = None
            
            for row_data in rows_to_write:
                if needs_category:
                    row_cat = str(row_data[cat_idx])
                    if row_cat != current_category:
                        for c_idx in range(1, num_cols + 1):
                            cell = ws.cell(row=row_pointer, column=c_idx)
                            cell.fill, cell.alignment = cat_fill, group_alignment
                        ws.cell(row=row_pointer, column=1, value="--- " + group_prefix + " " + row_cat + " ---").font = cat_font
                        ws.row_dimensions[row_pointer].outlineLevel = 1
                        ws.row_dimensions[row_pointer].hidden = False
                        current_category = row_cat
                        row_pointer += 1
                        
                for c_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_pointer, column=c_idx, value=str(value) if value is not None else "")
                ws.row_dimensions[row_pointer].outlineLevel = 2 if needs_category else 1
                ws.row_dimensions[row_pointer].hidden = False
                row_pointer += 1
                
            ws.sheet_properties.outlinePr.summaryBelow = False
            
            for col_idx in range(1, num_cols + 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 25
                
        wb.save(OUTPUT_FILEPATH)
    except Exception:
        pass
